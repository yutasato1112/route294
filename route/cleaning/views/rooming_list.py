from django.shortcuts import render
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView
from ..utils.home_util import read_csv, processing_list, dist_room, room_person, room_char

import io
import re
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font,Side, Border
import jaconv
from django.conf import settings
import os
import datetime

class roomingListView(TemplateView):
    template_name = "rooming_list.html"

    def get(self, request, *args, **kwargs):
        name = request.GET.get('editor_name')
        context = {
            'editor_name': name,
        }
        return render(self.request, self.template_name, context)
    
    def post(self, request, *args, **kwargs):
        method = "POST"
        today_csv = request.FILES.get("today")
        tomorrow_csv = request.FILES.get("tomorrow")
        editor_name = request.POST.get("editor_name")
        excel_template_path = os.path.join(settings.BASE_DIR, "static", "excel_template", "rooming_list.xlsx")
        output_excel_path = os.path.join(settings.MEDIA_ROOT, "rooming_list_modifies.xlsx")
        
        #csv読み込み
        today_data = read_csv_from_wincal(today_csv)
        tomorrow_data = read_csv_from_wincal(tomorrow_csv)
        room_info_data, times_by_time_data, master_key_data = read_csv()
        room_number_list = rooms = [x[0] for x in room_info_data]
        
        #連泊部屋リスト取得
        multiple_room_list = find_multiple_room_numbers(today_data, tomorrow_data)
        
        #未使用部屋リスト取得
        unuse_room_list = find_unuse_room_numbers(today_data, room_number_list)
        
        #Excel操作
        output_path = operating_excel(output_excel_path, multiple_room_list, unuse_room_list, excel_template_path, room_number_list)
        print('Finished Excel operation:', output_path)

        download_url = request.build_absolute_uri(
            os.path.join(settings.MEDIA_URL, os.path.basename(output_path))
        )
        
        #home遷移用
        from_report = False  
        #csv読み込み
        room_info_data, times_by_time_data, master_key_data = read_csv()
        
        #部屋を階別に二次元配列へ加工
        room_num_table = processing_list(room_info_data)
        
        #部屋をタイプ別に一次元配列に加工
        single_room_list, twin_room_list = dist_room(room_info_data)
        combined_rooms = []
        for i in range(len(room_num_table)):
            floor_data = []
            for j in range(len(room_num_table[i])):
                floor_data.append({
                    'room': room_num_table[i][j],
                    'status': '',
                })
            combined_rooms.append(floor_data)
            
        for floor in combined_rooms:
            for room_data in floor:
                if room_data['room'] in unuse_room_list:
                    room_data['status'] = "0"

        #日付取得
        #起動時刻が18時から24時の場合、翌日の日付を表示
        #それ以外は当日の日付を表示
        if datetime.datetime.now().hour >= 18:
            today = datetime.date.today() + datetime.timedelta(days=1)
        else:
            today = datetime.date.today()
        
        #連泊部屋入力欄対応
        padded_rooms = [''] * 100
        multiple_rows = [padded_rooms[i:i+10] for i in range(0, 100, 10)]      
        # multiple_rooms の値を上から順に埋める
        index = 0
        for r in range(10):
            for c in range(10):
                if index < len(multiple_room_list):
                    multiple_rows[r][c] = multiple_room_list[index]
                    index += 1
        

        context = {
            "download_url": download_url,
            'multiple_room_list': multiple_room_list,
            'unuse_room_list': unuse_room_list,
            'editor_name': editor_name,
        }
        return render(self.request, self.template_name, context)
        
        
        
def read_csv_from_wincal(file):
    # ファイルをテキストとして扱う（Shift-JIS で読みたい場合）
    wrapped = io.TextIOWrapper(file.file, encoding='shift_jis', errors='ignore')

    # DataFrame 読み込み
    data = (
        pd.read_csv(wrapped, header=None, usecols=[0, 10, 11, 15])
          .sort_values(by=10)
    )

    # 部屋番号の正規化関数
    def _normalize_room(val):
        if pd.isna(val):
            return val
        s = str(val).strip()

        if re.fullmatch(r"\d+\.0", s):
            try:
                s = str(int(float(s)))
            except Exception:
                pass

        # 先頭0除去（例: "0123" → "123"）
        if len(s) == 4 and s.startswith('0'):
            return s[1:]

        return s

    # 10列目が存在する場合のみ正規化
    if 10 in data.columns:
        data[10] = data[10].apply(_normalize_room)

    return data

def find_multiple_room_numbers(today_data, tomorrow_data):
    def _normalize_name(val):
        if pd.isna(val):
            return ''
        return str(val).strip()

    multiple_room_list = []
    for index, today_row in today_data.iterrows():
        today_room = today_row[10]
        today_reservation = today_row[0]
        today_name = _normalize_name(today_row[11])
        for index, tomorrow_row in tomorrow_data.iterrows():
            if today_room == tomorrow_row[10]:  # 部屋番号が一致するか確認
                if today_reservation == tomorrow_row[0]:  # 予約番号の一致も確認
                    if today_name == _normalize_name(tomorrow_row[11]):  # 名前の一致も確認
                        multiple_room_list.append(today_room)
    return multiple_room_list

def find_unuse_room_numbers(today_data,room_number_list):
    unuse_room_list = room_number_list.copy()
    for index, today_row in today_data.iterrows():
        room_number = today_row[10]
        if room_number in unuse_room_list:
            unuse_room_list.remove(room_number)
    return unuse_room_list

def operating_excel(output_excel_path,multiple_room_list, unuse_room_list, excel_template_path, room_number_list):
    #Excelテンプレート読み込み
    wb = openpyxl.load_workbook(excel_template_path)
    ws = wb.active

    # 背景色（黄色）
    yellow_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")

    # セル斜線（×）
    thin = Side(border_style="thin", color="000000")
    diagonal_slash = Border(
        diagonal=thin,
        diagonalUp=True,
        diagonalDown=False
    )
    
    # 全セルを走査、部分一致＆全角半角統一で判断
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            cell_str = str(cell.value)

            # セルの文字も全角→半角へ正規化
            cell_str_norm = jaconv.z2h(cell_str, digit=True)

            if any(room in cell_str_norm for room in multiple_room_list):
                cell.fill = yellow_fill

            if any(room in cell_str_norm for room in unuse_room_list):
                cell.border = diagonal_slash
    #各階連泊数集計
    multiple_count = count_multiple_rooms_by_floor(room_number_list, multiple_room_list)

    #各階清掃部屋数集計
    use_room_list = count_available_rooms_by_floor(room_number_list, unuse_room_list)
    
    # 集計結果をExcelに書き込み
    write_counts_to_excel(wb, multiple_count, use_room_list, output_excel_path)

    # 修正済み Excel を保存
    wb.save(output_excel_path)
    
    return output_excel_path



def write_counts_to_excel(wb, multiple_count, use_room_list, output_excel_path):
    ws = wb.active
    # 連泊数集計の書き込み
    base = '連泊：'
    for floor, count in multiple_count.items():
        if floor == '2F':
            content = base + str(count)
            ws['N40'] = content
        elif floor == '3F':
            content = base + str(count)
            ws['N28'] = content
        elif floor == '4F': 
            content = base + str(count)
            ws['N16'] = content
        elif floor == '5F':
            content = base + str(count)
            ws['N4'] = content
        elif floor == '6F':
            content = base + str(count)
            ws['AW52'] = content
        elif floor == '7F':
            content = base + str(count)
            ws['AW40'] = content
        elif floor == '8F': 
            content = base + str(count)
            ws['AW28'] = content
        elif floor == '9F':
            content = base + str(count)
            ws['AW16'] = content
        elif floor == '10F':
            content = base + str(count)
            ws['AW4'] = content
    
    # 清掃部屋数集計の書き込み
    base_front = '合計：'
    base_behind = '/17'
    for floor, count in use_room_list.items():
        if floor == '2F':
            content = base_front + str(count) + base_behind
            ws['Z40'] = content
        elif floor == '3F':
            content = base_front + str(count) + base_behind
            ws['Z28'] = content
        elif floor == '4F': 
            content = base_front + str(count) + base_behind
            ws['Z16'] = content
        elif floor == '5F':
            content = base_front + str(count) + base_behind
            ws['Z4'] = content
        elif floor == '6F':
            content = base_front + str(count) + base_behind
            ws['BI52'] = content
        elif floor == '7F':
            content = base_front + str(count) + base_behind
            ws['BI40'] = content
        elif floor == '8F': 
            content = base_front + str(count) + base_behind
            ws['BI28'] = content
        elif floor == '9F':
            content = base_front + str(count) + base_behind
            ws['BI16'] = content
        elif floor == '10F':
            content = base_front + str(count) + base_behind
            ws['BI4'] = content
                    
    #総清掃数
    total_cleaning = sum(use_room_list.values())
    ws['F57'] = total_cleaning
    #総連泊数
    total_multiple = sum(multiple_count.values())
    ws['F59'] = total_multiple

    return output_excel_path


def extract_floor_label(room_num):
    if len(room_num) == 3:
        return f"{int(room_num[0])}F"
    elif len(room_num) == 4:
        return f"{int(room_num[:2])}F"
    return None


def count_multiple_rooms_by_floor(room_number_list, multiple_room_list):
    # ===========================
    # 全角→半角統一
    # ===========================
    room_number_list = [jaconv.z2h(r, digit=True) for r in room_number_list]
    multiple_room_list = [jaconv.z2h(r, digit=True) for r in multiple_room_list]

    # ===========================
    # ① room_number_list から階層一覧を抽出
    # ===========================
    floors = set()

    for room in room_number_list:
        m = re.search(r"(\d{3,4})", room)
        if not m:
            continue
        num = m.group(1)
        floor_label = extract_floor_label(num)
        if floor_label:
            floors.add(floor_label)

    # 昇順ソート（1F,2F,...,10F,11F,...）
    sorted_floors = sorted(floors, key=lambda x: int(x[:-1]))

    # ===========================
    # ② multiple_room_list を階ごとにカウント
    # ===========================
    result = {f: 0 for f in sorted_floors}

    for room in multiple_room_list:
        m = re.search(r"(\d{3,4})", room)
        if not m:
            continue
        num = m.group(1)
        floor_label = extract_floor_label(num)
        if floor_label in result:
            result[floor_label] += 1

    return result

def count_available_rooms_by_floor(room_number_list, unuse_room_list):
    """
    room_number_list から unuse_room_list を除いた部屋を対象に、
    { '2F':部屋数, '3F':部屋数, '10F':部屋数, ... } を返す。
    """
    # ---------------------------
    # 全角→半角統一
    # ---------------------------
    room_number_list = [jaconv.z2h(r, digit=True) for r in room_number_list]
    unuse_room_list  = [jaconv.z2h(r, digit=True) for r in unuse_room_list]

    # ---------------------------
    # ① room_number_list → 階層一覧（全階）
    # ---------------------------
    floors = set()
    for room in room_number_list:
        m = re.search(r"(\d{3,4})", room)
        if not m:
            continue
        num = m.group(1)
        floors.add(extract_floor_label(num))

    # 階をソート（1F,2F,...,10F,...）
    sorted_floors = sorted(floors, key=lambda x: int(x[:-1]))

    # ---------------------------
    # ② 使用可能部屋を抽出
    # ---------------------------
    available_rooms = []
    for room in room_number_list:
        # unuse_room_list に「部分一致」でも含まれていれば除外
        if any(unuse in room for unuse in unuse_room_list):
            continue
        available_rooms.append(room)

    # ---------------------------
    # ③ 階別に部屋数カウント（0埋め）
    # ---------------------------
    result = {floor: 0 for floor in sorted_floors}

    for room in available_rooms:
        m = re.search(r"(\d{3,4})", room)
        if not m:
            continue
        num = m.group(1)
        floor_label = extract_floor_label(num)
        if floor_label in result:
            result[floor_label] += 1

    return result
